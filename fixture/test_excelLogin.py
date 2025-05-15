import pytest
from selenium.webdriver.common.by import By
import read_config
from selenium import webdriver
import time
import excelReader

@pytest.mark.parametrize("username, password",excelReader.get_data("ExcelFiles/loginData.xlsx","Sheet1"))
class TestExcel:
    def test_validlogin_usingExcel(self, username, password):
        self.driver = webdriver.Chrome()
   
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        self.driver.get("https://www.demoblaze.com/index.html")

        self.driver.find_element(By.ID,"login2").click()
        self.driver.find_element(By.ID,"loginusername").send_keys(username)
        time.sleep(3)
        self.driver.find_element(By.ID,"loginpassword").send_keys(password)
        time.sleep(3)
        self.driver.find_element(By.XPATH,"//button[text()='Log in']").click()
        time.sleep(5)
        if username == "admin":
            actual = self.driver.find_element(By.ID,"nameofuser").text
            assert username in actual
        else:
            print(self.driver.switch_to.alert.text)
            assert self.driver.switch_to.alert.text.__eq__("Wrong password.")
        self.driver.quit()


import openpyxl
def get_data(path , sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_row = sheet.max_row
    total_col = sheet.max_column

    for r in range(2 , total_row+1):
        row_list = []
        for c in range(1 , total_col+1):
            row_list.append(sheet.cell(r,c).value)
        final_list.append(row_list)
    return final_list